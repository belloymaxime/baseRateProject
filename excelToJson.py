import json
import openpyxl


path = "base rate assumptions.xlsx"

wb_obj = openpyxl.load_workbook(path, data_only=True)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
 
cell_obj = sheet_obj.cell(row=1, column=1)
 
sectors = ["All", "Consumer", "Digital Health", "Cloud Infrastructure", "Cybersecurity", "Fintech", "SaaS", "Supply Chain + Automation"]

rounds = [ "Pre Seed", "Seed", "Series A", "Series B", "Series C", "Series D", "Series E", "Series F", "Private Equity"]

data = {}

base = 3

for i in range(len(sectors)):
    data.update({sectors[i]: {}})

for i in range(len(sectors)):
    for k in range(len(rounds)):
        j = k+1
        data[sectors[i]].update({rounds[k]: {
            "Next round %": sheet_obj.cell(row=base, column=j+1).value,
            "Exit round %": sheet_obj.cell(row=base+1, column=j+1).value,
            "Exit < 200M %": sheet_obj.cell(row=base+2, column=j+1).value,
            "Exit 200M-500M %": sheet_obj.cell(row=base+3, column=j+1).value,
            "Exit 500M-1B %": sheet_obj.cell(row=base+4, column=j+1).value,
            "Exit 1B-2B %": sheet_obj.cell(row=base+5, column=j+1).value,
            "Exit 2B-5B %": sheet_obj.cell(row=base+6, column=j+1).value,
            "Exit 5B-10B %": sheet_obj.cell(row=base+7, column=j+1).value,
            "Exit > 10 B %": sheet_obj.cell(row=base+8, column=j+1).value,
            "Exit no Valuation %": sheet_obj.cell(row=base+9, column=j+1).value,
            "Median headcount": sheet_obj.cell(row=base+10, column=j+1).value,
            "Estimated ARR per FTE from Headcount + ICONIQ report": sheet_obj.cell(row=base+11, column=j+1).value,
            "Estimated ARR": sheet_obj.cell(row=base+12, column=j+1).value,
            "Average funding": sheet_obj.cell(row=base+13, column=j+1).value,
            "Median funding": sheet_obj.cell(row=base+14, column=j+1).value,
            "Average time to next round in typical series (days)": sheet_obj.cell(row=base+15, column=j+1).value,
            "Median time to next round in typical series(days)": sheet_obj.cell(row=base+16, column=j+1).value,
            "Average time to any next round (days)": sheet_obj.cell(row=base+17, column=j+1).value,
            "Median time to any next round (days)": sheet_obj.cell(row=base+18, column=j+1).value,
            "average time to exit (days)": sheet_obj.cell(row=base+19, column=j+1).value,
            "Median time to exit (days)": sheet_obj.cell(row=base+20, column=j+1).value,
            "Expected value of outcome": sheet_obj.cell(row=26, column=j+1).value
        }})
    if(base == 3):
        base = 37
    else:
        base += 25


with open("jsonVals.json", 'w') as json_file:
    json.dump(data, json_file, indent=4)

